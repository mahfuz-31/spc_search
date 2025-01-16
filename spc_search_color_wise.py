import requests as rq
from bs4 import BeautifulSoup as bs
import pandas as pd

def convert_to_number(s):
    # Remove commas from the string
    s = s.replace(",", "")
    # Convert to float if there's a decimal point, otherwise to convert_to_number
    return float(s) if '.' in s else int(s)

print('-------------- Mahfuz Special Search ----------------\n')
ordersdf = pd.read_excel('orders.xlsx')

df = pd.DataFrame()
columns = ["Buyer", "Order", 'Style',	'UoF',	'F. Color',	'G. Color', 'Y. Type', 'F. Type', 'GSM', 'Dia',	'G/F Order With S.Note Qty', 'G/F S.Note Qty',	'Net Grey Receive Qty',	'G/F Rcv Balance Qty',	'F/F Order with S.Note Qty',	'F/F S.Note Qty',	'F/F Delv Qty',	'F/F Delv. Balance Qty',	'Replacement Delivery',	'F/F Excess Delv.Qty',	'Transfer To',	'Transfer From',	'Return Receive',	'Return Delivery',	'Dyeing Unit',	'Order Sheet Receive Date', 'Cut Plan Start Date',	'Cut Plan End Date']
df = df.reindex(columns=df.columns.tolist() + columns)

row_idx = 0
no_of_orders = len(ordersdf['FRS No.'])

for index, row in ordersdf.iterrows():
    order = row['FRS No.']
    print("Process completed", int((index + 1) * 100 / no_of_orders), "%  <---->", 'FRS:', order)
    # print("\nCalculating for order: ", order)
    url = 'http://192.168.13.253/mymun/Work%20Order/combineSearchResult.php?Welcome=7&GetOrderNO=' + str(order)
    response = rq.get(url)

    html_content = bs(response.content, 'html.parser')
    rows = []
    for row in html_content.find_all('tr'):
        row_data = [cell.get_text(strip=True) for cell in row.find_all('td')]
        rows.append(row_data)

    row_idx2 = row_idx
    si = 1
    for row in rows:
        if len(row) > 0 and row[0].isnumeric() and row[0] != '0':
            if int(row[0]) < si:
                break
            df.loc[row_idx, 'Buyer'] = rows[3][0]
            df.loc[row_idx, 'Order'] = rows[3][1]
            df.loc[row_idx, 'Style'] = rows[3][2]
            df.loc[row_idx, 'UoF'] = row[1]
            df.loc[row_idx, 'F. Color'] = row[2][0:row[2].find('(')]
            df.loc[row_idx, 'G. Color'] = row[3]
            df.loc[row_idx, 'Y. Type'] = row[4]
            df.loc[row_idx, 'F. Type'] = row[5]
            df.loc[row_idx, 'GSM'] = convert_to_number(row[6])
            df.loc[row_idx, 'Dia'] = row[7]
            df.loc[row_idx, 'G/F Order With S.Note Qty'] = convert_to_number(row[8])
            df.loc[row_idx, 'G/F S.Note Qty'] = convert_to_number(row[9])
            df.loc[row_idx, 'Net Grey Receive Qty'] = convert_to_number(row[13])
            df.loc[row_idx, 'G/F Rcv Balance Qty'] = convert_to_number(row[15])
            row_idx += 1
            si += 1
    
    row_idx = row_idx2
    si = 0
    for row in rows:
        if len(row) > 0 and row[0] == 'SL NO.':
            si += 1
        if len(row) > 0 and si == 2 and row[0].isnumeric() and row[0] != '0':
            df.loc[row_idx, 'F/F Order with S.Note Qty'] = convert_to_number(row[8])
            df.loc[row_idx, 'F/F S.Note Qty'] = convert_to_number(row[9])
            df.loc[row_idx, 'F/F Delv Qty'] = convert_to_number(row[12])
            df.loc[row_idx, 'F/F Delv. Balance Qty'] = convert_to_number(row[15])
            df.loc[row_idx, 'Replacement Delivery'] = convert_to_number(row[16])
            df.loc[row_idx, 'F/F Excess Delv.Qty'] = convert_to_number(row[18])
            df.loc[row_idx, 'Transfer To'] = row[19]
            df.loc[row_idx, 'Transfer From'] = row[20]
            df.loc[row_idx, 'Return Receive'] = convert_to_number(row[22])
            df.loc[row_idx, 'Return Delivery'] = convert_to_number(row[23])
            df.loc[row_idx, 'Dyeing Unit'] = row[26]
            df.loc[row_idx, 'Order Sheet Receive Date'] = rows[3][4]
            df.loc[row_idx, 'Cut Plan Start Date'] = rows[3][5]
            df.loc[row_idx, 'Cut Plan End Date'] = rows[3][6]
            row_idx += 1

df.to_excel('output.xlsx', index=False)
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment

wb = load_workbook('output.xlsx')

ws = wb['Sheet1']

border = Border(left=Side(style='thin', color='0c8748'),
                right=Side(style='thin', color='0c8748'),
                top=Side(style='thin', color='0c8748'),
                bottom=Side(style='thin', color='0c8748'))

align = Alignment(horizontal='center', vertical='center', wrap_text=True)

for row in ws.iter_rows():
    for cell in row:
        cell.border = border
        cell.alignment = align
wb.save('color_wise_output.xlsx')

print("\nSuccessfully completed :)")